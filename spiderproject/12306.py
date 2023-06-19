# 登陆页面人工扫码登陆
import re
import time

import requests
import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains

driver = webdriver.Chrome()


class TrainSpider:
    login_url = 'https://kyfw.12306.cn/otn/resources/login.html'  # 登陆的连接
    profile_url = 'https://kyfw.12306.cn/otn/view/index.html'  # 个人主页的url
    left_ticket = 'https://kyfw.12306.cn/otn/leftTicket/init?linktypeid=dc'  # 余票查询的URL
    comfirm_url = 'https://kyfw.12306.cn/otn/confirmPassenger/initDc'  # 确认乘车人的订单

    def __init__(self, from_station, to_station, train_date, trains, passenger_lst):
        self.from_station = from_station
        self.to_station = to_station
        self.train_date = train_date
        self.station_code = self.init_station_code()  # 调用本类中初始化车站代号的方法
        self.trains = trains
        self.passenger_lst = passenger_lst

    # 初始化车站代码的方法
    def init_station_code(self):
        wb = openpyxl.load_workbook('车站名称及代号.xlsx')
        sheet = wb['station_name']
        lst = []
        for row in sheet.rows:
            sub_lst = []
            for cell in row:
                sub_lst.append(cell.value)
            lst.append(sub_lst)
        return dict(lst)

    # 余票查询
    def search_ticket(self):
        driver.get(self.left_ticket)
        # 出发站
        from_station_input = driver.find_element(by=By.ID, value='fromStation')
        # 到达站
        to_station_input = driver.find_element(by=By.ID, value='toStation')
        # 出发时间日期
        train_date_input = driver.find_element(by=By.ID, value='train_date')
        # 根据站点名称获取站点名称
        from_station_code = self.station_code.get(self.from_station)
        to_station_code = self.station_code.get(self.to_station)

        # 将站点代号赋值到input标签中
        driver.execute_script('arguments[0].value="%s"' % from_station_code, from_station_input)
        driver.execute_script('arguments[0].value="%s"' % to_station_code, to_station_input)
        driver.execute_script('arguments[0].value="%s"' % self.train_date, train_date_input)

        query_ticket_tag = driver.find_element(by=By.ID, value='query_ticket')
        query_ticket_tag.click()  # 点击查询余票的查询按钮
        print('余票查询成功')

        # 解析车次
        # 使用显示等待
        WebDriverWait(driver, 1000).until(
            EC.presence_of_element_located((By.XPATH, '//tbody[@id="queryLeftTable"]/tr'))
        )
        # 筛选出有数据的tr标签，去掉属性为datatran的tr
        trains = driver.find_elements(by=By.XPATH, value='//tbody[@id="queryLeftTable"]/tr[not(@datatran)]')

        is_flag = False  # 表示是否有余票
        # 遍历余票
        for train in trains:
            # print(train.text)
            infos = train.text.replace('\n', ' ').split(' ')
            train_no = infos[0]  # 车次
            if train_no in self.trains:  # self.trains表示的是要购买的车次  self.trains是字典类型
                # 得到座席
                seat_types = self.trains.get(train_no)  # 根据车次这个key得到的座席列表
                # 遍历座席一铺
                for seat_type in seat_types:
                    # 车次信息中索引为10的是二等座
                    count = infos[9]  # count为余票的张数，可能是具体的整数也可能是“有”
                    print('二等座:', count)
                    count = infos[8]  # 索引为9的是一等座
                    print('一等座:', count)
                    if seat_type == 'O':
                        count = infos[9]  # 获取余票的张数
                        if count.isdigit() or count == '有':
                            is_flag = True  # 表示有余票
                            break
                    elif seat_type == 'M':
                        count = infos[8]  # 获取一等座座席的张数
                        if count.isdigit() or count == '有':
                            is_flag = True
                            break  # 退出内层循环
                if is_flag:  # 如果有余票的情况
                    self.selected_no = train_no
                    order_btn = driver.find_element(by=By.XPATH, value='//a[@class="btn72"]')
                    order_btn.click()  # 单击进行预定
                    break  # 退出的是外层循环

    def confirm(self):
        WebDriverWait(driver, 1000).until(
            EC.url_to_be(self.comfirm_url)
        )
        WebDriverWait(driver, 1000).until(
            EC.presence_of_element_located((By.XPATH, '//ul[@id="normal_passenger_id"]/li/label'))
        )
        passangers = driver.find_elements(by=By.XPATH, value='//ul[@id="normal_passenger_id"]/li/label')

        for passanger in passangers:
            name = passanger.text
            if name in self.passenger_lst:
                passanger.click()

        # 确认坐席
        seat_select = Select(driver.find_element(by=By.ID, value='seatType_1'))
        # 从抢票人选择的车次中获取坐席的列表
        seat_types = self.trains.get(self.selected_no)

        for seat_type in seat_types:
            try:
                seat_select.select_by_value(seat_type)
            except NoSuchElementException:
                continue
            else:
                break
        # 提交订单
        submit_btn = driver.find_element(by=By.ID, value='submitOrder_id')
        submit_btn.click()
        print('成功提交车次订单')
        # 显示等待，判断核对以下信息对话框出现并确认按钮可以进行点击
        WebDriverWait(driver, 1000).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'dhtmlx_wins_body_inner'))
        )
        # 显示等待，确认按钮的出现
        WebDriverWait(driver, 1000).until(
            EC.presence_of_element_located((By.ID, 'qr_submit_id'))
        )
        # 定位确认按钮
        submit_button = driver.find_element(by=By.ID, value='qr_submit_id')
        while submit_button:  # 暴力点击
            try:
                submit_button.click()
                submit_button = driver.find_element(by=By.ID, value='qr_submit_id')
            except NoSuchElementException:
                break
        print('抢票成功')

    def login(self):
        # 打开登陆的页面
        driver.get(self.login_url)
        # 填写登陆的账号密码
        driver.find_element(by=By.ID, value='J-userName').send_keys('登陆账号')
        driver.find_element(by=By.ID, value='J-password').send_keys('登陆密码')

        driver.find_element(by=By.ID, value='J-login').click()

        # 滑块的验证
        # 等待滑块的出现
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.ID, 'nc_1_n1z')
            )
        )

        # 实例化行为链
        action = ActionChains(driver)
        action.click_and_hold(element)

        time.sleep(2)
        for i in range(10):
            action.move_by_offset(36, 0).perform()  # perform()立即执行动作链操作，move_by_offset(x, y):x水平方向  y竖直方向

        # 释放行为链
        action.release()

        # 显示等待，等待URL变成个人中心的URL，来判断登陆是否成功
        WebDriverWait(driver, 1000).until(
            EC.url_to_be(self.profile_url)
        )
        # 获取提示窗口，并点击确认的按钮
        # driver.find_element(by=By.CLASS_NAME, value='btn-primary').click()
        # 登陆成功
        print('登陆成功')

    # 封装一个抢票的功能
    def run(self):
        # 调用登陆
        self.login()
        # 调用余票查询的方法
        self.search_ticket()

        # 订单确认的方法
        self.confirm()


def satrt():
    spider = TrainSpider('北京', '深圳', '2023-06-22', {'G335': ['M', 'O']}, ['your_name'])  # M 一等座，O 二等座
    spider.run()


# 爬取城市名称和代号
# def get_city():
#     url = 'https://www.12306.cn/index/script/core/common/station_name_new_v10007.js'
#     header = {
#         'User - Agent': 'Mozilla / 5.0(Macintosh;IntelMacOSX10_15_7) AppleWebKit / 537.36(KHTML, likeGecko) Chrome / 113.0.0.0Safari / 537.36'
#     }
#     resp = requests.get(url, headers=header)
#     pattern = '([\u4e00-\u9fa5]+)\|([A-Z]+)'
#     station = re.findall(pattern, resp.text)
#     return station
#
#
# def save(lst):
#     wb = openpyxl.Workbook()
#     sheet = wb.create_sheet('station_name')
#     for item in lst:
#         sheet.append(item)
#     wb.save('车站名称及代号.xlsx')


if __name__ == '__main__':
    # lst = get_city()
    # save(lst)
    satrt()
